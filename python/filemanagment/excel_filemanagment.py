import os
from openpyxl import Workbook, load_workbook

def write_excel(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(["Name", "Age", "City"])

    ws.append(["Alice", 30, "Pune"])
    ws.append(["Bob", 25, "Mumbai"])

    wb.save(filename)
    print("Data written successfully.")

def read_excel(filename):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            print(row)
    else:
        print("File not found.")

def delete_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
        print("File deleted successfully.")
    else:
        print("File not found.")

#
# write_excel("myfile.xlsx")
# read_excel("myfile.xlsx")
delete_excel("myfile.xlsx")
